from __future__ import annotations
import csv
import io
from pathlib import Path
from datetime import datetime

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import DATA_DIR
from db import queries
from db.paths import test_type_folder_name

router = APIRouter()


@router.get("/session/{session_id}/csv")
async def download_merged_csv(session_id: int):
    session = await queries.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    merged_path = session.get("merged_path")
    if not merged_path or not Path(merged_path).exists():
        raise HTTPException(status_code=404, detail="No merged CSV available for this session")

    filename = Path(merged_path).name
    return FileResponse(
        path=merged_path,
        media_type="text/csv",
        filename=filename,
    )


@router.get("/all")
async def export_all():
    rows = await queries.export_all_sessions()
    if not rows:
        raise HTTPException(status_code=404, detail="No sessions found")

    # Build CSV in memory
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=gait_study_export.csv"},
    )


# ── Summary.xlsx — one sheet per test type + All Sessions ─────────────────

# Columns for per-test-type sheets (match the user's existing Summary.xlsx).
SUMMARY_COLUMNS = [
    ("Order",           "order"),
    ("New?",            "is_new_participant"),
    ("ID",              "participant_id"),
    ("Date",            "session_date"),
    ("injuries?",       "injuries"),
    ("Gender",          "gender"),
    ("Age",             "age"),
    ("UK shoe",         "shoe_size"),
    ("Insole size",     "insole_size"),
    ("status",          "status"),
    ("notes on setup",  "notes_setup"),
    ("notes on use",    "notes_use"),
    ("notes on data",   "notes_data"),
    ("general notes",   "notes"),
    ("condition order", "condition_order"),
    ("tempo direction", "tempo_direction"),
    ("weight direction","weight_direction"),
    ("agency aggregate","agency_aggregate"),
    ("ueqs pragmatic",  "ueqs_pragmatic"),
    ("ari immersion",   "ari_immersion"),
    ("cond1 actual",    "cond1_actual"),
    ("cond1 guess",     "cond1_guess"),
    ("cond1 correct",   "cond1_correct"),
    ("cond2 actual",    "cond2_actual"),
    ("cond2 guess",     "cond2_guess"),
    ("cond2 correct",   "cond2_correct"),
    ("cal file 1",      "cal_file_1_path"),
    ("cal file 2",      "cal_file_2_path"),
    ("cal file 3",      "cal_file_3_path"),
]

# "All Sessions" sheet prepends the test type column.
ALL_COLUMNS = [("Test Type", "test_type")] + SUMMARY_COLUMNS

_COL_WIDTHS = [8, 6, 8, 12, 18, 10, 6, 9, 11, 12, 28, 28, 28, 22, 16, 18, 18,
               16, 16, 16, 14, 14, 14, 14, 14, 14, 40, 40, 40]
_ALL_COL_WIDTHS = [24] + _COL_WIDTHS


def _write_sheet(ws, title: str, columns, rows_data, header_font, header_fill, widths):
    ws.cell(row=1, column=2, value=title).font = Font(bold=True, size=14)
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
    for row_idx, r in enumerate(rows_data, start=4):
        for col_idx, (_, key) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=r.get(key))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = w


async def _build_summary_workbook():
    """Workbook: first sheet = All Sessions; then one sheet per test type."""
    wb = Workbook()
    wb.remove(wb.active)

    test_types = await queries.list_test_types(include_archived=True)
    rows = await queries.export_all_sessions()

    by_test: dict[str, list[dict]] = {}
    for r in rows:
        key = r.get("test_type") or "Untitled"
        by_test.setdefault(key, []).append(r)

    sheet_order: list[str] = [t["name"] for t in test_types]
    for k in by_test:
        if k not in sheet_order:
            sheet_order.append(k)
    if not sheet_order:
        sheet_order = ["Sessions"]

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")

    # ── All Sessions (first sheet) ──────────────────────────────────────────
    ws_all = wb.create_sheet(title="All Sessions")
    _write_sheet(ws_all, "All Sessions", ALL_COLUMNS, rows, header_font, header_fill, _ALL_COL_WIDTHS)

    # ── Per test type ───────────────────────────────────────────────────────
    for tname in sheet_order:
        sheet_title = test_type_folder_name(tname)[:31]
        ws = wb.create_sheet(title=sheet_title)
        _write_sheet(ws, tname, SUMMARY_COLUMNS, by_test.get(tname, []),
                     header_font, header_fill, _COL_WIDTHS)

    return wb, by_test, sheet_order


async def _save_per_test_xlsx(by_test: dict, sheet_order: list[str]):
    """Write a summary xlsx into each test type's data folder."""
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")

    for tname in sheet_order:
        sessions_for_test = by_test.get(tname)
        if not sessions_for_test:
            continue
        folder_name = test_type_folder_name(tname)
        folder = DATA_DIR / folder_name
        folder.mkdir(parents=True, exist_ok=True)
        out_path = folder / f"{folder_name}_summary.xlsx"

        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title=folder_name[:31])
        _write_sheet(ws, tname, SUMMARY_COLUMNS, sessions_for_test,
                     header_font, header_fill, _COL_WIDTHS)
        wb.save(out_path)


@router.get("/summary.xlsx")
async def export_summary_xlsx():
    """Generate Summary.xlsx (All Sessions + one sheet per test type).

    Saves to data/Summary.xlsx and to data/<TestType>/<TestType>_summary.xlsx
    for each test type, then streams the combined file as a download.
    """
    wb, by_test, sheet_order = await _build_summary_workbook()

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    on_disk = DATA_DIR / "Summary.xlsx"
    wb.save(on_disk)

    await _save_per_test_xlsx(by_test, sheet_order)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Summary.xlsx"},
    )
